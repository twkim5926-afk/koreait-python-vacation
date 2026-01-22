# 파이썬에서 엑셀다루기 -> openpyxl
import openpyxl

# .xlsx 파일 생성(워크북생성)
wb = openpyxl.Workbook()

# 기본 시트를 현재 시트로 지정
ws = wb.active
ws.title = "첫번째시트" # 시트명 변경

# 엑셀좌표로 데이터 입력
ws["A1"] = "이름"
ws["B1"] = "나이"
ws["C1"] = "직업"

# row - 위에서 몇번째, column - 왼쪽에서 몇번째
ws.cell(row=2, column=1, value="홍길동")
ws.cell(row=2, column=2, value="20")
ws.cell(row=2, column=3, value="사무직")

wb.save("sample01.xlsx")

# 리스트에 있는 데이터를 한줄에 다 입력
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2. title = "학생명"

header = ["번호", "이름", "학년", "전공"]
ws.append(header)
ws2.append([1, "김철수", 2, "전자공학"])
ws2.append([2, "박민수", 2, "생명공학"])

wb2.save("sample02.xlsx")

